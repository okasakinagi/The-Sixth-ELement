"""
Analytics Service — 数据分析模块的业务逻辑层
"""

import csv
import io
from datetime import timezone as dt_timezone

from django.utils import timezone

from survey_management.mapper.analytics_mapper import AnalyticsMapper


class AnalyticsError(Exception):
    def __init__(self, status, message):
        self.status = status
        self.message = message
        super().__init__(message)


def _iso(dt):
    """把 datetime 转为带 Z 后缀的 ISO 8601 字符串。"""
    if dt is None:
        return None
    return dt.astimezone(dt_timezone.utc).isoformat().replace("+00:00", "Z")


def _make_anonymous_id(nickname):
    """
    对昵称做中间字符遮盖：
    - 长度 <= 2：保留首字符，其余替换为 *
    - 长度 > 2：保留首尾各 1 字符，中间全部替换为 *
    示例：'copilot' → 'c*****t'，'小明' → '小*'，'张三丰' → '张*丰'
    """
    if not nickname:
        return "*"
    if len(nickname) <= 2:
        return nickname[0] + "*"
    return nickname[0] + "*" * (len(nickname) - 2) + nickname[-1]


class AnalyticsService:
    def __init__(self):
        self.mapper = AnalyticsMapper()

    # ─── 权限检查 ─────────────────────────────────────

    def _authorize(self, user, survey):
        """检查用户是否有权查看该问卷的分析数据（发布者或管理员）。"""
        if survey.owner_id == user.id:
            return
        if self.mapper.is_admin(user):
            return
        raise AnalyticsError(403, "no permission to view this survey's analytics")

    # ─── 总览 ─────────────────────────────────────────

    def get_summary(self, user, survey_id):
        survey = self.mapper.get_survey(survey_id)
        if not survey:
            raise AnalyticsError(404, "survey not found")
        self._authorize(user, survey)

        responses_count = self.mapper.get_responses_count(survey_id)
        target = survey.target or None

        # 完成率 = 已填写份数 / 目标份数
        if target:
            completion_rate = round(responses_count / target, 2)
        else:
            completion_rate = None

        return {
            "survey_id": str(survey.id),
            "title": survey.title,
            "published_at": _iso(survey.updated_at),
            "responses_count": responses_count,
            "target": target,
            "completion_rate": completion_rate,
            "average_duration_seconds": self.mapper.get_avg_duration(survey_id),
        }

    # ─── 单题统计 ─────────────────────────────────────

    def get_questions_stats(self, user, survey_id, text_page=1, text_page_size=20):
        survey = self.mapper.get_survey(survey_id)
        if not survey:
            raise AnalyticsError(404, "survey not found")
        self._authorize(user, survey)

        if not survey.active_questionnaire_id:
            return {"items": []}

        questions = self.mapper.get_questions_with_options(
            survey.active_questionnaire_id
        )
        responses_count = self.mapper.get_responses_count(survey_id)

        items = []
        for q in questions:
            qtype = (q.type or "").lower()
            item = {
                "question_id": f"q_{q.id}",
                "order_no": q.order_no,
                "title": q.title,
                "type": qtype,
                "options": None,
                "texts": None,
                "texts_total": None,
            }

            if qtype in ("single", "multi"):
                item["options"] = self._count_options(q, responses_count)

            elif qtype in ("text", "multi-text"):
                text_page_size = min(int(text_page_size), 50)
                raw_items, total = self.mapper.get_text_answers(
                    q.id, int(text_page), text_page_size
                )
                item["texts"] = self._format_text_answers(raw_items, qtype)
                item["texts_total"] = total

            items.append(item)

        return {"items": items}

    def _count_options(self, question, responses_count):
        """统计单选/多选题各选项的人数和百分比。

        优先使用预设的 QuestionOption 保持顺序；如果没有预设选项，
        则从实际答案中动态归集标签（向后兼容旧数据/跳过选项录入的问卷）。
        """
        opts = question._prefetched_options
        # 预设选项的有序标签列表（可能为空）
        predefined_labels = [opt.label for opt in opts]
        # 用 dict 维持插入顺序：先填入预设标签（count=0），答案计数时再累加
        counts: dict[str, int] = {label: 0 for label in predefined_labels}

        raw_answers = self.mapper.get_choice_answers(question.id)
        qtype = (question.type or "").lower()

        for ans in raw_answers:
            if qtype == "single":
                label = (ans.get("value_text") or "").strip()
                if label:
                    counts[label] = counts.get(label, 0) + 1
            elif qtype == "multi":
                selected = ans.get("value_json") or []
                if isinstance(selected, list):
                    for item in selected:
                        label = str(item).strip()
                        if label:
                            counts[label] = counts.get(label, 0) + 1

        total = responses_count or 1
        # 输出时：先按预设顺序，再输出仅在答案中出现的额外标签
        output_labels = predefined_labels + [
            lbl for lbl in counts if lbl not in set(predefined_labels)
        ]
        return [
            {
                "label": label,
                "count": counts.get(label, 0),
                "ratio": round(counts.get(label, 0) / total, 2),
            }
            for label in output_labels
        ]

    def _format_text_answers(self, raw_items, qtype):
        """格式化文本回答列表，生成匿名 ID。"""
        result = []
        for item in raw_items:
            value = (
                item["value_json"]
                if qtype == "multi-text"
                else (item["value_text"] or "")
            )
            result.append(
                {
                    "response_id": item["response_id"],
                    "anonymous_id": _make_anonymous_id(item["nickname"]),
                    "value": value,
                    "submitted_at": _iso(item["submitted_at"]),
                }
            )
        return result

    # ─── 导出 ─────────────────────────────────────────

    def export_csv(self, user, survey_id):
        """
        生成 CSV 文件内容（bytes）。
        使用 UTF-8 with BOM（\\ufeff）确保 Excel 正确显示中文。
        """
        survey, questions, rows = self._build_export_data(user, survey_id)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        content = "\ufeff" + output.getvalue()  # BOM 保证 Excel 中文不乱码
        return content.encode("utf-8"), survey.title

    def export_xlsx(self, user, survey_id):
        """
        生成 Excel (.xlsx) 文件内容（bytes）。
        使用 openpyxl，自动调整列宽，中文无需任何特殊编码处理。
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise AnalyticsError(500, "openpyxl not installed")

        survey, questions, rows = self._build_export_data(user, survey_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "答卷数据"

        header_fill = PatternFill(
            start_color="1E4FB4", end_color="1E4FB4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, cell_value in enumerate(rows[0], start=1):
            cell = ws.cell(row=1, column=col_idx, value=cell_value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for row_idx, row in enumerate(rows[1:], start=2):
            for col_idx, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 自动调整列宽（最宽 50，最窄 10）
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), survey.title

    def _build_export_data(self, user, survey_id):
        """构造导出所需的二维表。返回 (survey, questions, rows)。"""
        survey = self.mapper.get_survey(survey_id)
        if not survey:
            raise AnalyticsError(404, "survey not found")
        self._authorize(user, survey)

        questions = []
        if survey.active_questionnaire_id:
            questions = self.mapper.get_questions_with_options(
                survey.active_questionnaire_id
            )

        responses = self.mapper.get_export_responses(survey_id)
        response_ids = [r.id for r in responses]
        all_answers = self.mapper.get_export_answers(response_ids)

        # 构建索引：response_id → {question_id → Answer}
        answer_index = {}
        for ans in all_answers:
            answer_index.setdefault(ans.response_id, {})[ans.question_id] = ans

        # 表头
        header = ["答卷ID", "匿名用户标识", "提交时间", "作答时长(秒)"]
        for q in questions:
            header.append(f"[Q{q.order_no}] {q.title}")

        rows = [header]
        for resp in responses:
            qtype_lower = {}
            for q in questions:
                qtype_lower[q.id] = (q.type or "").lower()

            row = [
                f"r_{resp.id}",
                _make_anonymous_id(resp.user.nickname if resp.user else ""),
                _iso(resp.submitted_at),
                resp.duration_seconds if resp.duration_seconds is not None else "",
            ]
            resp_answers = answer_index.get(resp.id, {})
            for q in questions:
                ans = resp_answers.get(q.id)
                if not ans:
                    row.append("")
                    continue
                qtype = qtype_lower[q.id]
                if qtype in ("multi", "multi-text"):
                    val = ans.value_json or []
                    row.append(
                        "，".join(str(v) for v in val)
                        if isinstance(val, list)
                        else str(val)
                    )
                else:
                    row.append(ans.value_text or "")
            rows.append(row)

        return survey, questions, rows
